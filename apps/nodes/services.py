"""节点创建/恢复与批量 Excel 导入相关业务逻辑。"""
import io
import ipaddress
import re
from typing import Any, Dict, List, Optional, Tuple

from django.core.exceptions import ObjectDoesNotExist
from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from apps.credentials.models import Credential
from utils.setting_service import get_setting

from .models import Node, NodeGroup


def mark_node_probe_success(node: Node) -> Node:
    """标记节点探测成功：置 online 并记录上次探测成功时间（不自动 save）。"""
    node.status = "online"
    node.last_probe_at = timezone.now()
    return node


def apply_nginx_probe_result(node: Node, success: bool, version: str = "") -> Node:
    """
    统一写入 Nginx 探测结果（不改 SSH status，不自动 save）。
    探测失败时清空版本并触发绑定 orphan；成功则标记可用。
    版本统一存纯数字（如 1.31.2），去掉 nginx/ 前缀。
    """
    from apps.releases.task_result import strip_nginx_version

    node.last_nginx_probe_at = timezone.now()
    if success:
        node.nginx_available = True
        if version:
            node.nginx_version = strip_nginx_version(version)
    else:
        node.nginx_available = False
        node.nginx_version = ""
        _orphan_bindings_after_nginx_missing(node)
    return node


def _orphan_bindings_after_nginx_missing(node: Node) -> None:
    """Nginx 确认不可用时，将该节点绑定标为远程已删除。"""
    try:
        from apps.configs.services import mark_node_bindings_orphaned

        mark_node_bindings_orphaned(node)
    except Exception:
        # 探测路径不应因配置侧异常中断 SSH 状态回写
        import logging

        logging.getLogger(__name__).exception(
            "节点 %s Nginx 缺失后 orphan 绑定失败", getattr(node, "hostname", node.pk)
        )


def nginx_ops_gate_message(node: Node) -> Optional[str]:
    """返回禁止 Nginx 业务操作的原因；允许时返回 None。"""
    if node.status != "online":
        return f"节点 {node.hostname} 非在线状态"
    if node.nginx_available is not True:
        if node.nginx_available is False:
            return f"节点 {node.hostname} 未检测到 Nginx，请先安装"
        return f"节点 {node.hostname} 尚未探测 Nginx，请先测试连接或采集版本"
    return None


def install_gate_message(node: Node) -> Optional[str]:
    """返回禁止安装的原因；允许时返回 None（仅校验 SSH 在线）。"""
    if node.status != "online":
        return f"节点 {node.hostname} 非在线状态"
    return None


def _get_node_credential(node):
    """返回节点关联的 SSH 凭证。"""
    return node.credential


# Excel 表头（须与模板完全一致）
IMPORT_HEADERS = (
    "主机名",
    "IP",
    "SSH端口",
    "所属环境",
    "Nginx路径",
    "Nginx主配置路径",
    "节点组",
    "凭证",
    "备注",
)
# 空值占位（视为未填）
_EMPTY_MARKERS = {"", "-", "—", "－", "无", "n/a", "na", "none"}

# 环境列归一化映射
_ENV_ALIASES = {
    "dev": "dev",
    "development": "dev",
    "开发": "dev",
    "开发环境": "dev",
    "test": "test",
    "testing": "test",
    "测试": "test",
    "测试环境": "test",
    "prod": "prod",
    "production": "prod",
    "生产": "prod",
    "生产环境": "prod",
}

# 导出时写短中文，便于再导入
_ENV_EXPORT_LABELS = {
    "dev": "开发",
    "test": "测试",
    "prod": "生产",
}


def _default_ssh_port() -> int:
    """读取系统设置中的默认 SSH 端口。"""
    try:
        return int(get_setting("node.ssh_default_port", "22") or 22)
    except (TypeError, ValueError):
        return 22


def _default_nginx_bin() -> str:
    """读取系统设置中的默认 Nginx 可执行文件路径。"""
    return (
        get_setting("config.default_nginx_bin", "/usr/sbin/nginx") or "/usr/sbin/nginx"
    ).strip()


def _default_nginx_conf() -> str:
    """读取系统设置中的默认 nginx.conf 主配置路径。"""
    from apps.configs.services import default_nginx_conf_path

    return (default_nginx_conf_path() or "/etc/nginx/nginx.conf").strip()


def _cell_str(value: Any) -> str:
    """将单元格值规范为去首尾空白的字符串。"""
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _is_empty_optional(value: str) -> bool:
    """判断可选字段是否视为空（含 - 占位）。"""
    return value.strip().lower() in _EMPTY_MARKERS


def _split_group_names(raw: str) -> List[str]:
    """按逗号/顿号/分号拆分节点组名称并去重保序。"""
    if _is_empty_optional(raw):
        return []
    parts = re.split(r"[,，、;；]", raw)
    names = []
    seen = set()
    for part in parts:
        name = part.strip()
        if not name or _is_empty_optional(name):
            continue
        if name in seen:
            continue
        seen.add(name)
        names.append(name)
    return names


def _normalize_environment(raw: str) -> Tuple[Optional[str], Optional[str]]:
    """归一化所属环境；空则默认 test。返回 (env_code, error)。"""
    if _is_empty_optional(raw):
        return "test", None
    key = raw.strip().lower()
    # 中文别名用原文 strip 再查
    mapped = _ENV_ALIASES.get(key) or _ENV_ALIASES.get(raw.strip())
    if mapped:
        return mapped, None
    return None, f"所属环境「{raw}」无法识别，请填写开发/测试/生产或 dev/test/prod"


def _parse_port(raw: str) -> Tuple[Optional[int], Optional[str]]:
    """解析必填 SSH 端口。"""
    if not raw or _is_empty_optional(raw):
        return None, "SSH端口不能为空"
    try:
        port = int(float(raw)) if "." in raw else int(raw)
    except (TypeError, ValueError):
        return None, f"SSH端口「{raw}」不是有效整数"
    if port < 1 or port > 65535:
        return None, f"SSH端口「{port}」超出范围（1-65535）"
    return port, None


def resolve_credential_by_name(
    name: str, user
) -> Tuple[Optional[Credential], Optional[str]]:
    """
    按名称解析已启用凭证。
    优先当前用户创建的同名凭证；否则若全局仅一条启用记录则采用；多条冲突返回错误。
    """
    if _is_empty_optional(name):
        return None, None
    qs = Credential.objects.filter(name=name, is_enabled=True)
    own = qs.filter(created_by=user).first()
    if own:
        return own, None
    matches = list(qs[:5])
    if not matches:
        return None, f"凭证「{name}」不存在或未启用"
    if len(matches) > 1:
        return None, f"凭证「{name}」存在多条启用记录，请改用唯一名称或由本人创建"
    return matches[0], None


def create_or_restore_node(
    user,
    *,
    hostname: str,
    ip: str,
    port: Optional[int] = None,
    credential=None,
    groups=None,
    environment: str = "dev",
    nginx_path: str = "",
    description: str = "",
) -> Tuple[Node, bool]:
    """
    创建节点；若同 IP 存在逻辑删除记录则恢复并覆盖字段。
    返回 (node, restored)。
    """
    if port is None:
        port = _default_ssh_port()
    if groups is None:
        groups = []
    environment = environment or "dev"

    deleted_node = Node.all_objects.filter(ip=ip, is_deleted=True).first()
    if deleted_node:
        deleted_node.hostname = hostname
        deleted_node.port = port
        deleted_node.credential = credential
        deleted_node.environment = environment
        deleted_node.nginx_path = nginx_path or ""
        deleted_node.description = description or ""
        deleted_node.status = "unknown"
        deleted_node.is_deleted = False
        deleted_node.deleted_at = None
        deleted_node.deleted_by = None
        deleted_node.save()
        deleted_node.groups.set(groups)
        return deleted_node, True

    node = Node(
        hostname=hostname,
        ip=ip,
        port=port,
        credential=credential,
        environment=environment,
        nginx_path=nginx_path or "",
        description=description or "",
        status="unknown",
        created_by=user,
    )
    node.save()
    if groups:
        node.groups.set(groups)
    return node, False


def build_node_export_bytes(nodes) -> bytes:
    """按导入表头将节点列表导出为 xlsx 字节内容。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "节点导出"
    ws.append(list(IMPORT_HEADERS))

    for node in nodes:
        group_names = "，".join(g.name for g in node.groups.all())
        cred_name = ""
        if node.credential_id and node.credential is not None:
            cred_name = node.credential.name or ""
        main_conf = ""
        try:
            setting = node.config_sync_setting
        except ObjectDoesNotExist:
            setting = None
        if setting is not None:
            main_conf = setting.main_conf_path or ""
        ws.append(
            [
                node.hostname or "",
                str(node.ip or ""),
                node.port if node.port is not None else "",
                _ENV_EXPORT_LABELS.get(node.environment, node.environment or ""),
                node.nginx_path or "",
                main_conf,
                group_names,
                cred_name,
                node.description or "",
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_import_template_bytes() -> bytes:
    """生成批量导入 Excel 模板字节内容。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "节点导入"
    ws.append(list(IMPORT_HEADERS))
    default_port = _default_ssh_port()
    default_bin = _default_nginx_bin()
    default_conf = _default_nginx_conf()
    # 示例行：可选列用 - 表示空（将套用默认规则）
    ws.append(
        [
            "web01",
            "10.10.10.101",
            default_port,
            "-",
            "-",
            "-",
            "-",
            "-",
            "-",
        ]
    )
    ws.append(
        [
            "web02",
            "10.10.10.102",
            default_port,
            "测试",
            default_bin,
            default_conf,
            "默认组",
            "default",
            "示例备注",
        ]
    )

    tip = wb.create_sheet("填写说明")
    tip.append(["说明"])
    tip.append(
        [
            "1. 表头必须为：主机名、IP、SSH端口、所属环境、Nginx路径、"
            "Nginx主配置路径、节点组、凭证、备注，请勿修改顺序或名称"
        ]
    )
    tip.append(["2. 主机名、IP、SSH端口为必填"])
    tip.append(["3. 所属环境可填开发/测试/生产或 dev/test/prod；空或 - 默认为测试环境"])
    tip.append(
        [f"4. Nginx路径为空或 - 时使用系统设置「默认 Nginx 可执行文件路径」（当前：{default_bin}）"]
    )
    tip.append(
        [
            "5. Nginx主配置路径映射配置同步的主配置文件路径；"
            f"空或 - 时使用系统设置「默认 nginx 主配置路径」（当前：{default_conf}）"
        ]
    )
    tip.append(
        ["6. 多个节点组可用逗号、顿号或分号分隔，最多 3 个，须为系统中已存在的组名"]
    )
    tip.append(["7. 凭证须填写已启用凭证的名称；节点组/凭证空或 - 表示不设置"])
    tip.append(["8. 备注可空；同 IP 若曾逻辑删除，导入将恢复原节点并关联历史记录"])
    tip.append(["9. 任一行校验失败则整批不导入"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_node_import_workbook(
    file_obj,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    解析上传的 xlsx，返回 (rows, errors)。
    rows 元素含 Excel 行号及各列原始字符串。
    """
    errors: List[Dict[str, Any]] = []
    try:
        wb = load_workbook(file_obj, read_only=True, data_only=True)
    except Exception:
        return [], [{"row": 0, "message": "无法解析 Excel 文件，请上传有效的 .xlsx"}]

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        wb.close()
        return [], [{"row": 1, "message": "表格为空，缺少表头"}]

    col_count = len(IMPORT_HEADERS)
    header_cells = [_cell_str(c) for c in (header or ())[:col_count]]
    while len(header_cells) < col_count:
        header_cells.append("")
    if tuple(header_cells) != IMPORT_HEADERS:
        wb.close()
        return [], [
            {
                "row": 1,
                "message": (
                    f"表头不正确，期望「{' / '.join(IMPORT_HEADERS)}」，"
                    f"实际「{' / '.join(header_cells)}」"
                ),
            }
        ]

    rows: List[Dict[str, Any]] = []
    for idx, raw in enumerate(rows_iter, start=2):
        cells = list(raw or ())
        while len(cells) < col_count:
            cells.append(None)
        values = [_cell_str(cells[i]) for i in range(col_count)]
        # 整行空白跳过
        if all(_is_empty_optional(v) for v in values):
            continue
        rows.append(
            {
                "row": idx,
                "hostname": values[0],
                "ip": values[1],
                "port_raw": values[2],
                "environment_raw": values[3],
                "nginx_path_raw": values[4],
                "main_conf_path_raw": values[5],
                "groups_raw": values[6],
                "credential_raw": values[7],
                "description_raw": values[8],
            }
        )
    wb.close()

    if not rows:
        errors.append({"row": 0, "message": "没有可导入的数据行"})
    return rows, errors


def validate_node_import_rows(
    rows: List[Dict[str, Any]], user
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    全量校验导入行；任一行失败则 cleaned 为空。
    返回 (cleaned, errors)。
    """
    errors: List[Dict[str, Any]] = []
    if not rows:
        return [], [{"row": 0, "message": "没有可导入的数据行"}]

    hostname_seen: Dict[str, int] = {}
    ip_seen: Dict[str, int] = {}
    pending: List[Dict[str, Any]] = []
    default_bin = _default_nginx_bin()
    default_conf = _default_nginx_conf()

    for item in rows:
        row_no = item["row"]
        hostname = item["hostname"]
        ip = item["ip"]
        row_errors: List[str] = []

        if not hostname:
            row_errors.append("主机名不能为空")
        elif len(hostname) > 100:
            row_errors.append("主机名长度不能超过 100")

        if not ip:
            row_errors.append("IP 不能为空")
        else:
            try:
                ipaddress.ip_address(ip)
            except ValueError:
                row_errors.append(f"IP「{ip}」格式不合法")

        port, port_err = _parse_port(item.get("port_raw") or "")
        if port_err:
            row_errors.append(port_err)

        environment, env_err = _normalize_environment(
            item.get("environment_raw") or ""
        )
        if env_err:
            row_errors.append(env_err)

        nginx_raw = item.get("nginx_path_raw") or ""
        if _is_empty_optional(nginx_raw):
            nginx_path = default_bin
        else:
            nginx_path = nginx_raw
            if len(nginx_path) > 255:
                row_errors.append("Nginx路径长度不能超过 255")

        conf_raw = item.get("main_conf_path_raw") or ""
        if _is_empty_optional(conf_raw):
            main_conf_path = default_conf
        else:
            main_conf_path = conf_raw
            if len(main_conf_path) > 500:
                row_errors.append("Nginx主配置路径长度不能超过 500")

        desc_raw = item.get("description_raw") or ""
        description = "" if _is_empty_optional(desc_raw) else desc_raw

        if hostname:
            if hostname in hostname_seen:
                row_errors.append(
                    f"文件内主机名「{hostname}」与第 {hostname_seen[hostname]} 行重复"
                )
            else:
                hostname_seen[hostname] = row_no

        if ip:
            if ip in ip_seen:
                row_errors.append(f"文件内 IP「{ip}」与第 {ip_seen[ip]} 行重复")
            else:
                ip_seen[ip] = row_no

        group_names = _split_group_names(item.get("groups_raw") or "")
        groups = []
        if len(group_names) > 3:
            row_errors.append("节点最多只能关联 3 个节点组")
        else:
            for gname in group_names:
                try:
                    groups.append(NodeGroup.objects.get(name=gname))
                except NodeGroup.DoesNotExist:
                    row_errors.append(f"节点组「{gname}」不存在")

        credential, cred_err = resolve_credential_by_name(
            item.get("credential_raw") or "", user
        )
        if cred_err:
            row_errors.append(cred_err)

        if row_errors:
            for msg in row_errors:
                errors.append({"row": row_no, "message": msg})
            continue

        pending.append(
            {
                "row": row_no,
                "hostname": hostname,
                "ip": ip,
                "port": port,
                "environment": environment,
                "nginx_path": nginx_path,
                "main_conf_path": main_conf_path,
                "description": description,
                "groups": groups,
                "credential": credential,
            }
        )

    if pending:
        hostnames = [p["hostname"] for p in pending]
        ips = [p["ip"] for p in pending]
        active_by_hostname = {
            n.hostname: n
            for n in Node.objects.filter(hostname__in=hostnames).only(
                "id", "hostname", "ip"
            )
        }
        active_by_ip = {
            n.ip: n
            for n in Node.objects.filter(ip__in=ips).only("id", "hostname", "ip")
        }
        for p in pending:
            if p["hostname"] in active_by_hostname:
                other = active_by_hostname[p["hostname"]]
                errors.append(
                    {
                        "row": p["row"],
                        "message": (
                            f"主机名「{p['hostname']}」已被活跃节点占用"
                            f"（IP {other.ip}）"
                        ),
                    }
                )
            if p["ip"] in active_by_ip:
                other = active_by_ip[p["ip"]]
                errors.append(
                    {
                        "row": p["row"],
                        "message": (
                            f"IP「{p['ip']}」已被活跃节点「{other.hostname}」占用"
                        ),
                    }
                )

    if errors:
        return [], errors
    return pending, []


@transaction.atomic
def apply_node_import(cleaned: List[Dict[str, Any]], user) -> Dict[str, int]:
    """事务内写入已校验通过的节点行，并同步配置主路径。"""
    from apps.configs.services import save_sync_path

    created = 0
    restored = 0
    for item in cleaned:
        node, was_restored = create_or_restore_node(
            user,
            hostname=item["hostname"],
            ip=item["ip"],
            port=item["port"],
            credential=item.get("credential"),
            groups=item.get("groups") or [],
            environment=item.get("environment") or "test",
            nginx_path=item.get("nginx_path") or "",
            description=item.get("description") or "",
        )
        save_sync_path(
            node,
            item.get("main_conf_path") or _default_nginx_conf(),
            user,
        )
        if was_restored:
            restored += 1
        else:
            created += 1
    return {"created": created, "restored": restored, "total": created + restored}


# ---------- SSH 测活 / 采集后台任务 ----------


def run_unlock_ssh_test(task_id, node_ids):
    """解锁后逐节点测试并写入标准结果树"""
    from apps.releases.models import TaskCenterTask
    from apps.releases.task_result import (
        build_tree_result,
        item_failed,
        item_success,
        node_header,
    )
    from utils.ssh import get_nginx_version, test_ssh_connection

    node_list = list(Node.objects.filter(id__in=node_ids).order_by("id"))
    total = len(node_list)
    TaskCenterTask.objects.filter(pk=task_id).update(
        status="running",
        progress=5,
        detail="正在解锁并测试连接...",
        started_at=timezone.now(),
    )
    success_count = 0
    fail_count = 0
    done = 0
    node_blocks = []

    for node in node_list:
        try:
            credential = _get_node_credential(node)
            if not credential:
                node.status = "unknown"
                node.save()
                fail_count += 1
                node_blocks.append(node_header(node.ip, node.hostname))
                node_blocks.append(item_failed("SSH连接", "未配置凭证"))
            elif not credential.is_enabled:
                node.status = "offline"
                node.save()
                fail_count += 1
                node_blocks.append(node_header(node.ip, node.hostname))
                node_blocks.append(item_failed("SSH连接", "凭证已禁用"))
            else:
                if credential.auth_type == "password":
                    success, message = test_ssh_connection(
                        node.ip,
                        node.port,
                        credential.username,
                        password=credential.get_password(),
                    )
                else:
                    success, message = test_ssh_connection(
                        node.ip,
                        node.port,
                        credential.username,
                        private_key=credential.get_private_key(),
                    )
                if success:
                    mark_node_probe_success(node)
                    nginx_path = node.nginx_path if node.nginx_path else None
                    version_success, version_info = get_nginx_version(
                        node.ip,
                        node.port,
                        credential.username,
                        password=credential.get_password()
                        if credential.auth_type == "password"
                        else None,
                        private_key=credential.get_private_key()
                        if credential.auth_type == "key"
                        else None,
                        nginx_path=nginx_path,
                    )
                    apply_nginx_probe_result(
                        node, version_success, version_info if version_success else ""
                    )
                    success_count += 1
                    node_blocks.append(node_header(node.ip, node.hostname))
                    node_blocks.append(item_success("SSH连接"))
                else:
                    node.status = "offline"
                    fail_count += 1
                    node_blocks.append(node_header(node.ip, node.hostname))
                    node_blocks.append(item_failed("SSH连接", message))
                node.save()
        except Exception as e:
            fail_count += 1
            node_blocks.append(node_header(node.ip, node.hostname))
            node_blocks.append(item_failed("SSH连接", f"异常: {str(e)}"))

        done += 1
        TaskCenterTask.objects.filter(pk=task_id).update(
            progress=int(done * 100 / total) if total else 100,
            detail=f"执行中：成功{success_count}，失败{fail_count}，已完成{done}/{total}",
            updated_at=timezone.now(),
        )

    status = "success" if fail_count == 0 else "failed"
    TaskCenterTask.objects.filter(pk=task_id).update(
        status=status,
        progress=100,
        finished_at=timezone.now(),
        detail=f"执行完成：成功{success_count}，失败{fail_count}，共{total}",
        result=build_tree_result(success_count, fail_count, total, node_blocks),
    )


def run_single_node_ssh_test(
    task_id,
    host,
    ssh_port,
    credential_id,
    target_hostname,
    node_id=None,
):
    """单节点 SSH 测试，结果写入标准树"""
    from apps.releases.models import TaskCenterTask
    from apps.releases.task_cancel import finish_if_active
    from apps.releases.task_result import (
        build_tree_result,
        item_failed,
        item_success,
        node_header,
    )
    from utils.ssh import get_nginx_version, test_ssh_connection

    credential = Credential.objects.get(pk=credential_id)
    _has_node = node_id is not None
    _node_id = node_id

    TaskCenterTask.objects.filter(pk=task_id).update(
        status="running",
        progress=5,
        detail="正在测试SSH连接...",
        started_at=timezone.now(),
    )

    try:
        if credential.auth_type == "password":
            success, message = test_ssh_connection(
                host, ssh_port, credential.username, password=credential.get_password()
            )
        else:
            success, message = test_ssh_connection(
                host,
                ssh_port,
                credential.username,
                private_key=credential.get_private_key(),
            )

        TaskCenterTask.objects.filter(pk=task_id).update(
            progress=60, detail="正在更新节点状态..."
        )

        if success and _has_node:
            _node = Node.objects.get(id=_node_id)
            mark_node_probe_success(_node)
            nginx_path = _node.nginx_path if _node.nginx_path else None
            version_success, version_info = get_nginx_version(
                host,
                ssh_port,
                credential.username,
                password=credential.get_password()
                if credential.auth_type == "password"
                else None,
                private_key=credential.get_private_key()
                if credential.auth_type == "key"
                else None,
                nginx_path=nginx_path,
            )
            apply_nginx_probe_result(
                _node, version_success, version_info if version_success else ""
            )
            _node.save()
        elif not success and _has_node:
            _node = Node.objects.get(id=_node_id)
            _node.status = "offline"
            _node.save()

        status = "success" if success else "failed"
        blocks = [node_header(host, target_hostname)]
        if success:
            blocks.append(item_success("SSH连接"))
        else:
            blocks.append(item_failed("SSH连接", message))
        finish_if_active(
            task_id,
            status=status,
            progress=100,
            finished_at=timezone.now(),
            detail=f"连接{'成功' if success else '失败'}",
            result=build_tree_result(
                1 if success else 0,
                0 if success else 1,
                1,
                blocks,
            ),
        )
    except Exception as e:
        blocks = [
            node_header(host, target_hostname),
            item_failed("SSH连接", str(e)),
        ]
        finish_if_active(
            task_id,
            status="failed",
            progress=100,
            finished_at=timezone.now(),
            detail=f"执行异常: {str(e)}",
            result=build_tree_result(0, 1, 1, blocks),
        )


def _batch_test_one_node(node):
    """批量测活中的单节点 SSH 测试，返回结果字典"""
    from utils.ssh import get_nginx_version, test_ssh_connection

    try:
        if node.is_locked:
            return {
                "node_id": node.id,
                "hostname": node.hostname,
                "ip": node.ip,
                "success": False,
                "message": "节点已锁定",
            }
        credential = _get_node_credential(node)
        if not credential:
            return {
                "node_id": node.id,
                "hostname": node.hostname,
                "ip": node.ip,
                "success": False,
                "message": "未配置凭证",
            }
        if not credential.is_enabled:
            return {
                "node_id": node.id,
                "hostname": node.hostname,
                "ip": node.ip,
                "success": False,
                "message": "关联凭证已禁用",
            }

        if credential.auth_type == "password":
            success, message = test_ssh_connection(
                node.ip,
                node.port,
                credential.username,
                password=credential.get_password(),
            )
        else:
            success, message = test_ssh_connection(
                node.ip,
                node.port,
                credential.username,
                private_key=credential.get_private_key(),
            )

        if success:
            mark_node_probe_success(node)
            nginx_path = node.nginx_path if node.nginx_path else None
            version_success, version_info = get_nginx_version(
                node.ip,
                node.port,
                credential.username,
                password=(
                    credential.get_password()
                    if credential.auth_type == "password"
                    else None
                ),
                private_key=(
                    credential.get_private_key()
                    if credential.auth_type == "key"
                    else None
                ),
                nginx_path=nginx_path,
            )
            apply_nginx_probe_result(
                node, version_success, version_info if version_success else ""
            )
        else:
            node.status = "offline"

        node.save()

        return {
            "node_id": node.id,
            "hostname": node.hostname,
            "ip": node.ip,
            "success": success,
            "message": message,
        }
    except Exception as e:
        return {
            "node_id": node.id,
            "hostname": node.hostname,
            "ip": node.ip,
            "success": False,
            "message": str(e),
        }


def run_batch_node_ssh_test(task_id, node_ids, max_workers):
    """批量 SSH 测试，结果写入标准树"""
    from concurrent.futures import ThreadPoolExecutor, as_completed

    from apps.releases.models import TaskCenterTask
    from apps.releases.task_cancel import finish_if_active, is_cancelled, update_if_active
    from apps.releases.task_result import (
        build_tree_result,
        item_failed,
        item_success,
        node_header,
    )

    test_nodes = list(Node.objects.filter(id__in=node_ids).order_by("id"))
    TaskCenterTask.objects.filter(pk=task_id).update(
        status="running",
        started_at=timezone.now(),
        progress=0,
        detail=f"执行中：0/{len(test_nodes)}",
    )

    success_count = 0
    fail_count = 0
    done = 0
    node_blocks = []

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_node = {
            executor.submit(_batch_test_one_node, node): node for node in test_nodes
        }
        for future in as_completed(future_to_node):
            if is_cancelled(task_id):
                break
            result = future.result()
            done += 1
            node_blocks.append(node_header(result.get("ip"), result.get("hostname")))
            if result.get("success"):
                success_count += 1
                node_blocks.append(item_success("SSH连接"))
            else:
                fail_count += 1
                node_blocks.append(item_failed("SSH连接", result.get("message", "")))

            update_if_active(
                task_id,
                progress=(
                    int(done * 100 / len(test_nodes)) if test_nodes else 100
                ),
                detail=f"执行中：成功 {success_count}，失败 {fail_count}，已完成 {done}/{len(test_nodes)}",
            )

    if is_cancelled(task_id):
        return
    status = "success" if fail_count == 0 else "failed"
    finish_if_active(
        task_id,
        status=status,
        progress=100,
        finished_at=timezone.now(),
        detail=f"执行完成：成功 {success_count}，失败 {fail_count}，共 {len(test_nodes)}",
        result=build_tree_result(
            success_count, fail_count, len(test_nodes), node_blocks
        ),
    )


def run_node_system_info_task(task_id, node_id, credential_id):
    """后台采集节点系统信息并回写状态"""
    import json

    from apps.releases.models import TaskCenterTask
    from utils.ssh import get_system_info

    node = Node.objects.get(id=node_id)
    credential = Credential.objects.get(pk=credential_id)
    TaskCenterTask.objects.filter(pk=task_id).update(
        status="running",
        progress=5,
        detail="正在采集系统信息...",
        started_at=timezone.now(),
    )

    try:
        if credential.auth_type == "password":
            success, system_info = get_system_info(
                node.ip,
                node.port,
                credential.username,
                password=credential.get_password(),
            )
        else:
            success, system_info = get_system_info(
                node.ip,
                node.port,
                credential.username,
                private_key=credential.get_private_key(),
            )

        if success:
            mark_node_probe_success(node)
            node.save()
        else:
            node.status = "offline"
            node.save()

        status = "success" if success else "failed"
        result_data = system_info if success else system_info
        TaskCenterTask.objects.filter(pk=task_id).update(
            status=status,
            progress=100,
            finished_at=timezone.now(),
            detail=f"系统信息采集{'成功' if success else '失败'}",
            result=json.dumps(system_info, ensure_ascii=False)
            if success
            else str(result_data),
        )
    except Exception as e:
        TaskCenterTask.objects.filter(pk=task_id).update(
            status="failed",
            progress=100,
            finished_at=timezone.now(),
            detail=f"执行异常: {str(e)}",
        )


def run_node_nginx_version_task(task_id, node_id, credential_id, nginx_path=None):
    """后台检测节点 Nginx 版本并回写"""
    from apps.releases.models import TaskCenterTask
    from utils.ssh import get_nginx_version

    node = Node.objects.get(id=node_id)
    credential = Credential.objects.get(pk=credential_id)
    TaskCenterTask.objects.filter(pk=task_id).update(
        status="running",
        progress=5,
        detail="正在检测 Nginx 版本...",
        started_at=timezone.now(),
    )

    try:
        if credential.auth_type == "password":
            success, output = get_nginx_version(
                node.ip,
                node.port,
                credential.username,
                password=credential.get_password(),
                nginx_path=nginx_path,
            )
        else:
            success, output = get_nginx_version(
                node.ip,
                node.port,
                credential.username,
                private_key=credential.get_private_key(),
                nginx_path=nginx_path,
            )

        if success:
            mark_node_probe_success(node)
            apply_nginx_probe_result(node, True, output)
            node.save(
                update_fields=[
                    "nginx_version",
                    "nginx_available",
                    "last_nginx_probe_at",
                    "status",
                    "last_probe_at",
                    "updated_at",
                ]
            )
        else:
            # SSH 已连通仅 nginx -v 失败时不改 status，避免误标离线
            apply_nginx_probe_result(node, False, "")
            node.save(
                update_fields=[
                    "nginx_version",
                    "nginx_available",
                    "last_nginx_probe_at",
                    "updated_at",
                ]
            )

        status = "success" if success else "failed"
        TaskCenterTask.objects.filter(pk=task_id).update(
            status=status,
            progress=100,
            finished_at=timezone.now(),
            detail=f"Nginx 版本检测{'成功: ' + output if success else '失败'}",
            result=output if success else output,
        )
    except Exception as e:
        TaskCenterTask.objects.filter(pk=task_id).update(
            status="failed",
            progress=100,
            finished_at=timezone.now(),
            detail=f"执行异常: {str(e)}",
        )
