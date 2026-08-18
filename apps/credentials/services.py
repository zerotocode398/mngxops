"""凭证启用测试、xlsx 导出与批量导入。"""
import io
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import StringIO
from typing import Any, Dict, List, Optional, Tuple

from django.db import close_old_connections, transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from apps.nodes.models import Node
from apps.nodes.services import apply_openssh_probe_result, mark_node_probe_success
from apps.releases.models import TaskCenterTask
from apps.releases.task_cancel import finish_if_active
from utils.setting_service import get_setting
from utils.ssh import get_openssh_version, test_ssh_connection

from .models import Credential, CredentialEnableTask

# 导出/导入表头须一致
EXPORT_HEADERS = (
    "名称",
    "SSH用户",
    "认证方式",
    "密码",
    "私钥",
    "是否启用",
    "描述",
)
IMPORT_HEADERS = EXPORT_HEADERS

_EMPTY_MARKERS = {"", "-", "—", "－", "无", "n/a", "na", "none"}

_AUTH_ALIASES = {
    "password": "password",
    "pwd": "password",
    "密码": "password",
    "密码认证": "password",
    "key": "key",
    "密钥": "key",
    "秘钥": "key",
    "密钥认证": "key",
    "秘钥认证": "key",
    "私钥": "key",
}

_ENABLED_TRUE = {"是", "启用", "已启用", "true", "1", "yes", "y", "on"}
_ENABLED_FALSE = {"否", "禁用", "已禁用", "false", "0", "no", "n", "off"}


def _cell_str(value: Any) -> str:
    """将单元格值规范为去首尾空白的字符串。"""
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _is_empty_optional(value: str) -> bool:
    """判断可选字段是否视为空（含 - 占位）。"""
    return (value or "").strip().lower() in _EMPTY_MARKERS


def is_valid_private_key(key_str: str) -> bool:
    """校验私钥是否为合法 RSA/DSA/ECDSA/Ed25519 格式。"""
    import paramiko

    if not (key_str or "").strip():
        return False
    key_types = [
        paramiko.RSAKey,
        paramiko.DSSKey,
        paramiko.ECDSAKey,
        paramiko.Ed25519Key,
    ]
    for key_type in key_types:
        try:
            key_type.from_private_key(StringIO(key_str))
            return True
        except Exception:
            continue
    return False


def build_credential_export_bytes(credentials) -> bytes:
    """将凭证列表导出为含明文密码/私钥的 xlsx 字节内容。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "凭证导出"
    ws.append(list(EXPORT_HEADERS))

    for cred in credentials:
        password = ""
        private_key = ""
        if cred.auth_type == "password":
            password = cred.get_password() or ""
        elif cred.auth_type == "key":
            private_key = cred.get_private_key() or ""
        ws.append(
            [
                cred.name or "",
                cred.username or "",
                cred.get_auth_type_display(),
                password,
                private_key,
                "是" if cred.is_enabled else "否",
                cred.description or "",
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_credential_import_template_bytes() -> bytes:
    """生成凭证批量导入 Excel 模板字节内容。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "凭证导入"
    ws.append(list(IMPORT_HEADERS))
    ws.append(
        [
            "default-password",
            "root",
            "密码认证",
            "ChangeMe123!",
            "",
            "是",
            "示例：密码认证",
        ]
    )
    ws.append(
        [
            "default-key",
            "root",
            "密钥认证",
            "",
            "-----BEGIN OPENSSH PRIVATE KEY-----\n...(粘贴完整私钥)...\n-----END OPENSSH PRIVATE KEY-----",
            "是",
            "示例：密钥认证（请替换为真实私钥）",
        ]
    )

    tip = wb.create_sheet("填写说明")
    tip.append(["说明"])
    tip.append(
        [
            "1. 表头必须为：名称、SSH用户、认证方式、密码、私钥、是否启用、描述，"
            "请勿修改顺序或名称"
        ]
    )
    tip.append(["2. 名称、SSH用户、认证方式为必填"])
    tip.append(
        ["3. 认证方式可填：密码认证/密码/password，或 密钥认证/密钥/key"]
    )
    tip.append(["4. 密码认证须填密码，私钥列可空；密钥认证须填合法私钥，密码列可空"])
    tip.append(["5. 是否启用可填 是/否（或启用/禁用）；空或 - 默认启用"])
    tip.append(
        [
            "6. 同名凭证（当前登录用户下）将更新；新名称则新建；"
            "任一行校验失败则整批不导入"
        ]
    )
    tip.append(["7. 私钥单元格可含换行；勿使用带口令的加密私钥"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_credential_import_workbook(
    file_obj,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """解析上传的凭证 xlsx，返回 (rows, errors)。"""
    errors: List[Dict[str, Any]] = []
    try:
        wb = load_workbook(file_obj, read_only=True, data_only=True)
    except Exception:
        return [], [{"row": 0, "message": "无法解析 Excel 文件，请上传有效的 .xlsx"}]

    try:
        ws = wb[wb.sheetnames[0]]
    except Exception:
        wb.close()
        return [], [{"row": 0, "message": "Excel 中没有可用工作表"}]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        wb.close()
        return [], [{"row": 1, "message": "表格为空，缺少表头"}]

    col_count = len(IMPORT_HEADERS)
    header_cells = [_cell_str(c) for c in (header or ())[:col_count]]
    while len(header_cells) < col_count:
        header_cells.append("")
    if tuple(header_cells) != IMPORT_HEADERS:
        wb.close()
        return [], [
            {
                "row": 1,
                "message": (
                    f"表头不正确，期望「{' / '.join(IMPORT_HEADERS)}」，"
                    f"实际「{' / '.join(header_cells)}」"
                ),
            }
        ]

    rows: List[Dict[str, Any]] = []
    for idx, raw in enumerate(rows_iter, start=2):
        cells = list(raw or ())
        while len(cells) < col_count:
            cells.append(None)
        # 私钥可能含换行，仅对非私钥列 strip；私钥保留内部换行仅去首尾
        name = _cell_str(cells[0])
        username = _cell_str(cells[1])
        auth_raw = _cell_str(cells[2])
        password = "" if cells[3] is None else str(cells[3])
        if isinstance(cells[3], float) and cells[3] == int(cells[3]):
            password = str(int(cells[3]))
        private_key = "" if cells[4] is None else str(cells[4])
        enabled_raw = _cell_str(cells[5])
        description = "" if cells[6] is None else str(cells[6]).strip()

        values_for_empty = [
            name,
            username,
            auth_raw,
            password.strip(),
            private_key.strip(),
            enabled_raw,
            description,
        ]
        if all(_is_empty_optional(v) for v in values_for_empty):
            continue

        rows.append(
            {
                "row": idx,
                "name": name,
                "username": username,
                "auth_raw": auth_raw,
                "password": password.strip("\n\r") if password else "",
                "private_key": private_key.strip() if private_key else "",
                "enabled_raw": enabled_raw,
                "description": (
                    "" if _is_empty_optional(description) else description
                ),
            }
        )
    wb.close()

    if not rows:
        errors.append({"row": 0, "message": "没有可导入的数据行"})
    return rows, errors


def _normalize_auth_type(raw: str) -> Tuple[Optional[str], Optional[str]]:
    """归一化认证方式；返回 (auth_type, error)。"""
    if _is_empty_optional(raw):
        return None, "认证方式不能为空"
    key = raw.strip().lower()
    mapped = _AUTH_ALIASES.get(key) or _AUTH_ALIASES.get(raw.strip())
    if mapped:
        return mapped, None
    return None, f"认证方式「{raw}」无法识别，请填写密码认证或密钥认证"


def _normalize_enabled(raw: str) -> Tuple[Optional[bool], Optional[str]]:
    """归一化是否启用；空则默认 True。"""
    if _is_empty_optional(raw):
        return True, None
    key = raw.strip().lower()
    if key in {x.lower() for x in _ENABLED_TRUE} or raw.strip() in _ENABLED_TRUE:
        return True, None
    if key in {x.lower() for x in _ENABLED_FALSE} or raw.strip() in _ENABLED_FALSE:
        return False, None
    return None, f"是否启用「{raw}」无法识别，请填写是/否"


def validate_credential_import_rows(
    rows: List[Dict[str, Any]], user
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """全量校验凭证导入行；任一行失败则 cleaned 为空。"""
    errors: List[Dict[str, Any]] = []
    if not rows:
        return [], [{"row": 0, "message": "没有可导入的数据行"}]

    name_seen: Dict[str, int] = {}
    pending: List[Dict[str, Any]] = []

    for item in rows:
        row_no = item["row"]
        name = (item.get("name") or "").strip()
        username = (item.get("username") or "").strip()
        row_errors: List[str] = []

        if not name:
            row_errors.append("名称不能为空")
        elif len(name) > 100:
            row_errors.append("名称长度不能超过 100")

        if not username:
            row_errors.append("SSH用户不能为空")
        elif len(username) > 100:
            row_errors.append("SSH用户长度不能超过 100")

        auth_type, auth_err = _normalize_auth_type(item.get("auth_raw") or "")
        if auth_err:
            row_errors.append(auth_err)

        is_enabled, en_err = _normalize_enabled(item.get("enabled_raw") or "")
        if en_err:
            row_errors.append(en_err)

        password = item.get("password") or ""
        private_key = item.get("private_key") or ""
        # 去掉示例占位私钥
        if private_key and "...(粘贴完整私钥)..." in private_key:
            private_key = ""

        if auth_type == "password":
            if not password.strip():
                row_errors.append("密码认证方式必须填写密码")
            private_key = ""
        elif auth_type == "key":
            if not private_key.strip():
                row_errors.append("密钥认证方式必须填写私钥")
            elif not is_valid_private_key(private_key):
                row_errors.append(
                    "私钥格式无效，请提供合法的 RSA/DSA/ECDSA/Ed25519 格式私钥"
                )
            password = ""

        if name:
            if name in name_seen:
                row_errors.append(
                    f"文件内名称「{name}」与第 {name_seen[name]} 行重复"
                )
            else:
                name_seen[name] = row_no

        if row_errors:
            for msg in row_errors:
                errors.append({"row": row_no, "message": msg})
            continue

        pending.append(
            {
                "row": row_no,
                "name": name,
                "username": username,
                "auth_type": auth_type,
                "password": password,
                "private_key": private_key,
                "is_enabled": is_enabled,
                "description": item.get("description") or "",
            }
        )

    if errors:
        return [], errors
    return pending, []


def apply_credential_import(cleaned: List[Dict[str, Any]], user) -> Dict[str, int]:
    """按校验结果写入凭证：同名（当前用户）更新，否则新建。"""
    created = 0
    updated = 0
    with transaction.atomic():
        for item in cleaned:
            existing = Credential.objects.filter(
                name=item["name"], created_by=user
            ).first()
            if existing:
                existing.username = item["username"]
                existing.auth_type = item["auth_type"]
                existing.is_enabled = item["is_enabled"]
                existing.description = item["description"]
                if item["auth_type"] == "password":
                    existing.password = item["password"]
                    existing.private_key = ""
                else:
                    existing.private_key = item["private_key"]
                    existing.password = ""
                existing.save()
                updated += 1
            else:
                cred = Credential(
                    name=item["name"],
                    username=item["username"],
                    auth_type=item["auth_type"],
                    password=item["password"] if item["auth_type"] == "password" else "",
                    private_key=(
                        item["private_key"] if item["auth_type"] == "key" else ""
                    ),
                    is_enabled=item["is_enabled"],
                    description=item["description"],
                    created_by=user,
                )
                cred.save()
                created += 1
    return {
        "created": created,
        "updated": updated,
        "total": created + updated,
    }


def _update_credential_test_result(credential, fail_count):
    """根据测试失败数更新凭证的最后测试结果字段"""
    credential.last_test_time = timezone.now()
    if fail_count == 0:
        credential.last_test_result = "success"
    elif fail_count >= credential.node_set.filter(is_locked=False).count():
        credential.last_test_result = "failed"
    else:
        credential.last_test_result = "partial"
    credential.save(update_fields=["last_test_time", "last_test_result"])


def _run_credential_enable_task(task_id, credential_id):
    """后台线程执行凭证启用后的关联节点批量连接测试"""
    # Ensure thread owns a clean DB connection.
    close_old_connections()

    try:
        from apps.releases.task_result import (
            build_tree_result,
            item_failed,
            item_success,
            node_header,
        )

        task = CredentialEnableTask.objects.get(pk=task_id)
        credential = Credential.objects.get(pk=credential_id)
        center_task_id = task.task_center_id
        cred_label = credential.name or f"凭证#{credential_id}"

        task.status = "running"
        task.started_at = timezone.now()
        task.save(update_fields=["status", "started_at", "updated_at"])

        if center_task_id:
            TaskCenterTask.objects.filter(pk=center_task_id).update(
                status="running",
                started_at=timezone.now(),
                progress=0,
                detail="测试开始",
            )

        nodes = list(Node.objects.filter(credential=credential, is_locked=False).order_by("id"))
        task.total_count = len(nodes)
        task.skipped_count = Node.objects.filter(
            credential=credential, is_locked=True
        ).count()
        task.save(update_fields=["total_count", "skipped_count", "updated_at"])

        if not nodes:
            task.status = "completed"
            task.finished_at = timezone.now()
            task.message = "无可测试节点"
            task.save(update_fields=["status", "finished_at", "message", "updated_at"])
            # 更新凭证的最后测试结果
            credential.last_test_time = timezone.now()
            credential.last_test_result = "unknown"
            credential.save(update_fields=["last_test_time", "last_test_result"])
            if center_task_id:
                TaskCenterTask.objects.filter(pk=center_task_id).update(
                    status="success",
                    progress=100,
                    finished_at=timezone.now(),
                    detail="无可测试节点",
                    result=f"执行完成：成功 0，失败 0，共 0\n凭证 {cred_label} 无可测试节点",
                )
            return

        max_workers = min(int(get_setting("node.batch_max_count", "3")), len(nodes))

        def _test_node(node):
            """对单个节点执行SSH连接测试，返回 (node, success, message)"""
            try:
                if credential.auth_type == "password":
                    success, message = test_ssh_connection(
                        node.ip,
                        node.port,
                        credential.username,
                        password=credential.get_password(),
                    )
                else:
                    success, message = test_ssh_connection(
                        node.ip,
                        node.port,
                        credential.username,
                        private_key=credential.get_private_key(),
                    )
                if success:
                    mark_node_probe_success(node)
                    ossh_success, ossh_info = get_openssh_version(
                        node.ip,
                        node.port,
                        credential.username,
                        password=credential.get_password()
                        if credential.auth_type == "password"
                        else None,
                        private_key=credential.get_private_key()
                        if credential.auth_type == "key"
                        else None,
                    )
                    apply_openssh_probe_result(
                        node, ossh_success, ossh_info if ossh_success else ""
                    )
                    node.save(
                        update_fields=[
                            "status",
                            "last_probe_at",
                            "openssh_version",
                            "last_openssh_probe_at",
                            "updated_at",
                        ]
                    )
                else:
                    node.status = "offline"
                    node.save(update_fields=["status", "updated_at"])
                return node, success, message or ("连接成功" if success else "连接失败")
            except Exception as exc:
                node.status = "offline"
                node.save(update_fields=["status", "updated_at"])
                return node, False, str(exc)

        success_count = 0
        fail_count = 0
        completed_count = 0
        node_blocks = []

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = [executor.submit(_test_node, n) for n in nodes]
            for future in as_completed(futures):
                completed_count += 1
                node, ok, message = future.result()
                if ok:
                    success_count += 1
                    node_blocks.append(node_header(node.ip, node.hostname))
                    node_blocks.append(item_success("SSH连接"))
                else:
                    fail_count += 1
                    node_blocks.append(node_header(node.ip, node.hostname))
                    node_blocks.append(item_failed("SSH连接", message))

                CredentialEnableTask.objects.filter(pk=task.pk).update(
                    completed_count=completed_count,
                    success_count=success_count,
                    failed_count=fail_count,
                    updated_at=timezone.now(),
                )
                if center_task_id:
                    TaskCenterTask.objects.filter(pk=center_task_id).update(
                        progress=int((completed_count / len(nodes)) * 100),
                        detail=f"成功 {success_count}，失败 {fail_count}，共 {len(nodes)}",
                        updated_at=timezone.now(),
                    )

        task.refresh_from_db()
        task.status = "completed"
        task.finished_at = timezone.now()
        skip_tip = f"，锁定跳过 {task.skipped_count}" if task.skipped_count else ""
        task.message = (
            f"凭证 {cred_label}：成功 {task.success_count}，失败 {task.failed_count}{skip_tip}"
        )
        task.save(
            update_fields=["status", "finished_at", "message", "updated_at"]
        )

        # 更新凭证的最后测试结果
        _update_credential_test_result(credential, fail_count)

        if center_task_id:
            result_text = build_tree_result(
                task.success_count, task.failed_count, len(nodes), node_blocks
            )
            if task.skipped_count:
                result_text += f"\n锁定跳过 {task.skipped_count} 台"
            finish_if_active(
                center_task_id,
                status="success" if task.failed_count == 0 else "failed",
                progress=100,
                finished_at=timezone.now(),
                result=result_text,
                detail=f"成功 {task.success_count} / 失败 {task.failed_count}",
                target_hostnames=",".join(n.hostname for n in nodes if n.hostname),
                target_ips=",".join(n.ip for n in nodes if n.ip),
                target_configs=cred_label,
            )
    except Exception as exc:
        CredentialEnableTask.objects.filter(pk=task_id).update(
            status="failed",
            finished_at=timezone.now(),
            message=f"任务失败: {exc}",
            updated_at=timezone.now(),
        )
        try:
            failed_task = CredentialEnableTask.objects.get(pk=task_id)
            if failed_task.task_center_id:
                finish_if_active(
                    failed_task.task_center_id,
                    status="failed",
                    finished_at=timezone.now(),
                    progress=100,
                    result=f"  [失败] 凭证测试 - 失败原因: {exc}",
                    detail=f"任务失败: {exc}",
                )
        except CredentialEnableTask.DoesNotExist:
            pass
    finally:
        close_old_connections()
